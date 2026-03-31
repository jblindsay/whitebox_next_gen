from pathlib import Path
import re
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import FormulaRule


ROOT = Path('/Users/johnlindsay/Documents/programming/Rust/whitebox_next_gen')
INPUT_TSV = ROOT / 'TOOL_INVENTORY_SPREADSHEET_FORMAT.tsv'
OUTPUT_XLSX = ROOT / 'TOOL_PORTING_PROGRESS.xlsx'

SECTION_RE = re.compile(
    r'^(AGRICULTURE TOOLS|DATA TOOLS|GEOMORPHOMETRY|GIS - RASTER OVERLAY & ANALYSIS|HYDROLOGY|REMOTE SENSING|LIDAR PROCESSING|STREAM NETWORK ANALYSIS|MATH & STATISTICAL TOOLS)'
)


def parse_summary(lines):
    rows = []
    in_summary = False
    header = []
    for ln in lines:
        if ln.strip() == 'CATEGORY SUMMARY TAB':
            in_summary = True
            continue
        if in_summary and ln.startswith('Category\t'):
            header = ln.split('\t')
            continue
        if in_summary:
            if ln.strip() == '---':
                break
            if ln.strip():
                parts = ln.split('\t')
                if len(parts) >= len(header):
                    rows.append(dict(zip(header, parts)))
    return rows


def parse_sections(lines):
    sections = {}
    i = 0
    while i < len(lines):
        ln = lines[i]
        m = SECTION_RE.match(ln)
        if not m:
            i += 1
            continue

        title = m.group(1)
        j = i + 1
        while j < len(lines) and (not lines[j].strip() or lines[j].strip() == '---'):
            j += 1
        if j >= len(lines):
            break

        header = lines[j].split('\t')
        rows = []
        k = j + 1
        while k < len(lines):
            row_ln = lines[k]
            if row_ln.strip() == '---' or SECTION_RE.match(row_ln):
                break
            if not row_ln.strip():
                k += 1
                continue
            if '\t' in row_ln:
                parts = row_ln.split('\t')
                if len(parts) < len(header):
                    parts += [''] * (len(header) - len(parts))
                rows.append(dict(zip(header, parts[:len(header)])))
            k += 1

        sections[title] = rows
        i = k

    return sections


def parse_int_prefix(value):
    m = re.search(r'\d+', str(value))
    return int(m.group(0)) if m else 0


def autosize_and_wrap(ws):
    for col in range(1, ws.max_column + 1):
        max_len = 0
        letter = get_column_letter(col)
        for row in range(1, ws.max_row + 1):
            val = ws.cell(row, col).value
            if val is None:
                continue
            max_len = max(max_len, len(str(val)))
        ws.column_dimensions[letter].width = min(max(12, max_len + 2), 60)

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.alignment = Alignment(vertical='top', wrap_text=True)


def add_status_conditional_formatting(ws, status_col, notes_col):
    red_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
    amber_fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
    green_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')

    max_row = ws.max_row
    if max_row < 2:
        return

    data_range = f'A2:G{max_row}'

    ws.conditional_formatting.add(
        data_range,
        FormulaRule(
            formula=[f'OR(${status_col}2="NOT_PORTED",ISNUMBER(SEARCH("exception",${notes_col}2)))'],
            fill=red_fill,
        ),
    )
    ws.conditional_formatting.add(
        data_range,
        FormulaRule(
            formula=[f'OR(${status_col}2="partial",${status_col}2="todo")'],
            fill=amber_fill,
        ),
    )
    ws.conditional_formatting.add(
        data_range,
        FormulaRule(
            formula=[f'${status_col}2="done"'],
            fill=green_fill,
        ),
    )


def main():
    lines = INPUT_TSV.read_text(encoding='utf-8').splitlines()
    summary_rows = parse_summary(lines)
    sections = parse_sections(lines)

    wb = Workbook()
    wb.remove(wb.active)

    header_fill = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True)

    ws_progress = wb.create_sheet('Overall Progress')
    progress_headers = [
        'Thematic Area',
        'Legacy Tool Count',
        'Ported Tool Count',
        'Remaining To Port',
        'Percent Complete',
        'Primary Tier',
        'Notes',
    ]
    ws_progress.append(progress_headers)
    for cell in ws_progress[1]:
        cell.fill = header_fill
        cell.font = header_font

    for row in summary_rows:
        cat = row.get('Category', '')
        if cat == 'TOTAL':
            continue

        legacy_raw = row.get('Legacy_Count', '0')
        ported_raw = row.get('Ported_Count', '0')
        remaining_raw = row.get('Remaining_To_Port', '0')

        legacy_n = parse_int_prefix(legacy_raw)
        ported_n = parse_int_prefix(ported_raw)
        pct = (ported_n / legacy_n * 100.0) if legacy_n else 0.0

        ws_progress.append([
            cat,
            legacy_raw,
            ported_raw,
            remaining_raw,
            round(pct, 1),
            row.get('Platform', ''),
            row.get('Notes', ''),
        ])

    for r in range(2, ws_progress.max_row + 1):
        ws_progress.cell(r, 5).number_format = '0.0"%"'
    ws_progress.auto_filter.ref = f'A1:G{ws_progress.max_row}'
    ws_progress.freeze_panes = 'A2'

    progress_red_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
    progress_green_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
    ws_progress.conditional_formatting.add(
        f'A2:G{ws_progress.max_row}',
        FormulaRule(formula=['$E2<100'], fill=progress_red_fill),
    )
    ws_progress.conditional_formatting.add(
        f'A2:G{ws_progress.max_row}',
        FormulaRule(formula=['$E2>=100'], fill=progress_green_fill),
    )

    area_order = [
        ('AGRICULTURE TOOLS', 'Agriculture'),
        ('DATA TOOLS', 'Data Tools'),
        ('GEOMORPHOMETRY', 'Geomorphometry'),
        ('GIS - RASTER OVERLAY & ANALYSIS', 'GIS'),
        ('HYDROLOGY', 'Hydrology'),
        ('REMOTE SENSING', 'Remote Sensing'),
        ('LIDAR PROCESSING', 'LiDAR Processing'),
        ('STREAM NETWORK ANALYSIS', 'Stream Network Analysis'),
        ('MATH & STATISTICAL TOOLS', 'Math & Statistics'),
    ]

    sheet_headers = [
        'New_Tool_Name',
        'Legacy_Tool_Name',
        'Port_Tier',
        'Port_Status',
        'Is_New_NonLegacy',
        'Notes',
        'Python_Wrapper_Status',
    ]

    for source_name, sheet_name in area_order:
        ws = wb.create_sheet(sheet_name)
        ws.append(sheet_headers)
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
        ws.freeze_panes = 'A2'

        for row in sections.get(source_name, []):
            tool = row.get('Tool_Name', '').strip()
            status = row.get('Status', '').strip()
            platform = row.get('Platform', '').strip()
            notes = row.get('Notes', '').strip()
            python_wrapper_status = row.get('Python_Wrapper_Status', '').strip()
            legacy_eq = row.get('Legacy_Equivalent', '').strip().lower()

            is_new = False
            legacy_name = tool

            notes_lower = notes.lower()
            if tool.startswith('NEW_ENHANCEMENT') or 'new_enhancement' in notes_lower or 'new - beyond legacy' in notes_lower:
                is_new = True
                legacy_name = ''
            elif legacy_eq in {'no', 'false'}:
                is_new = True
                legacy_name = ''
            elif status.upper() == 'NOT_PORTED' and ('legacy_exception' in platform.lower() or 'dormant' in platform.lower()):
                legacy_name = tool
            else:
                legacy_name = tool

            if status.upper() == 'NOT_PORTED' and ('legacy_exception' in platform.lower() or 'dormant' in platform.lower()):
                if notes:
                    notes += ' | '
                notes += 'Legacy exception; intentionally not ported.'

            if platform:
                tier_note = f'Tier: {platform}'
                if notes:
                    notes = f'{tier_note} | {notes}'
                else:
                    notes = tier_note

            ws.append([
                tool,
                legacy_name,
                platform,
                status,
                'Yes' if is_new else 'No',
                notes,
                python_wrapper_status,
            ])

        ws.auto_filter.ref = f'A1:F{ws.max_row}'
        add_status_conditional_formatting(ws, status_col='D', notes_col='F')

    for ws in wb.worksheets:
        autosize_and_wrap(ws)

    wb.save(OUTPUT_XLSX)
    print(f'Wrote: {OUTPUT_XLSX}')


if __name__ == '__main__':
    main()
